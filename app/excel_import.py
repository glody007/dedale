from openpyxl import load_workbook

'''extract data from excel file and retrieve
   list of dico of every row'''
class ExcelDataExtractor:

    def __init__(self, path):

        self.students = []

        wb = load_workbook(path)
        #selection de la feuille contenant les etudiants
        sheet = wb["students"]

        """nombre de lignes de la premiere ligne
           a la derniere ligne non vide"""
        max_row = sheet.max_row
        """nombre de colonnes de la premiere colonne
           a la derniere colonne non vide"""
        max_column = sheet.max_column

        #premiere cellule de datas
        first_cell = 'A2'
        #derniere cellule de datas
        last_cell  = 'D' + str(max_row)

        cells = sheet[first_cell : last_cell]

        for c1, c2, c3, c4 in cells:
            student = {'first_name' : c1.value,
                       'last_name' : c2.value,
                       'forename' : c3.value,
                       'sex' : c4.value}
            self.students.append(student)

    def getStudents(self):
        return self.students
