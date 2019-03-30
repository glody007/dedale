from openpyxl import Workbook, load_workbook
import unittest
import os
from datas import *
from .. import ExcelDataExtractor
from ...utils import *
from openpyxl.utils.cell import get_column_interval

#@ remember to add test for file with titles only
class ExcelImportTestCase(unittest.TestCase):
    def setUp(self):
        wb = Workbook()
        sheet = wb.active
        self.insert_rows_in_sheet(sheet)
        basedir = os.path.abspath(os.path.dirname(__file__))
        path_doc_with_datas = os.path.join(basedir, 'doc_with_datas.xlsx')
        wb.save(path_doc_with_datas)
        self.extractor_with_datas = ExcelDataExtractor(path_doc_with_datas)

        wb_without_datas = Workbook()
        path_doc_without_datas = os.path.join(basedir, 'doc_without_datas.xlsx')
        wb_without_datas.save(path_doc_without_datas)
        self.extractor_without_datas = ExcelDataExtractor(path_doc_without_datas)


    def insert_rows_in_sheet(self, sheet):
        empty_row = []
        test_titles = get_columns_titles()
        sheet.append(test_titles)
        for row in get_datas_by_rows():
            sheet.append(empty_row)
            sheet.append(row)

    def tearDown(self):
        pass

    def extracted_contains_test_datas(self, extracted_datas, test_datas):
        for i in range(len(extracted_datas)):
            if not list_contains_another(extracted_datas[i], test_datas[i]):
                return False
        return True

    def is_none(self, data):
        for value in data:
            if value is not None:
                return False
        return True

    def there_is_empty_row(self, extracted_datas):
        for data in extracted_datas:
            if self.is_none(data):
                return True
        return False

    def test_get_datas(self):
        extracted_datas = self.extractor_with_datas.get_datas()

        there_is_empty_row = self.there_is_empty_row(extracted_datas)
        self.assertFalse(there_is_empty_row, "extracted datas has empty row")

        test_datas = get_datas_by_rows()
        self.assertTrue(self.extracted_contains_test_datas(extracted_datas,
                                                           test_datas),
                        "extracted datas from xlsx doesn't contains test datas")

        no_extracted_datas = self.extractor_without_datas.get_datas()
        self.assertTrue(no_extracted_datas is None,
                        "extracted datas from empty xlsx is not none")

    def test_get_columns_titles(self):
        titles = self.extractor_with_datas.get_columns_titles()
        test_titles = get_columns_titles()
        extracted_titles_contains_test_titles =\
                                    list_contains_another(titles, test_titles)
        self.assertTrue(extracted_titles_contains_test_titles,
                      "extracted titles from xlsx doesn't contains test titles")

        no_extracted_titles = self.extractor_without_datas.get_columns_titles()
        self.assertTrue(no_extracted_titles is None,
                        "extracted titles from empty xlsx is not none")

    def labeled_datas_contains_test_datas(self, labeled_datas, labeled_test_datas):
        for i in range(len(labeled_test_datas)):
            contained = dict_contient(labeled_datas[i], labeled_test_datas[i])
            if not contained:
                return False
        return True

    def test_datas_labeled_by_columns_alias(self):
        columns_alias = get_columns_alias()
        labeled_datas = self.\
                        extractor_with_datas.\
                        get_datas_labeled_by_columns_alias(columns_alias)
        labeled_test_datas = get_labeled_datas()
        self.assertTrue(self.labeled_datas_contains_test_datas(labeled_datas,
                                                            labeled_test_datas),
                        "labeled datas doesn't contain labeled test datas")

        no_labeled_datas = self.\
                           extractor_without_datas.\
                           get_datas_labeled_by_columns_alias(columns_alias)
        self.assertEqual(no_labeled_datas, None,
                         "labeled datas is not none from xlsx without datas")

        alias_with_min_out = get_columns_alias_min_key_out_of_interval()
        with self.assertRaises(KeyError):
            self.extractor_with_datas.\
                 get_datas_labeled_by_columns_alias(alias_with_min_out)

        alias_with_max_out = get_columns_alias_min_key_out_of_interval()
        with self.assertRaises(KeyError):
            self.extractor_with_datas.\
                 get_datas_labeled_by_columns_alias(alias_with_max_out)

    def test_datas_labeled_by_titles_alias(self):
        titles_alias = get_titles_alias()
        labeled_datas = self.\
                        extractor_with_datas.\
                        get_datas_labeled_by_titles_alias(titles_alias)
        labeled_test_datas = get_labeled_datas()
        self.assertTrue(self.labeled_datas_contains_test_datas(labeled_datas,
                                                            labeled_test_datas),
                        "labeled datas doesn't contain labeled test datas")

        wrong_titles_alias = get_wrong_titles_alias()
        with self.assertRaises(KeyError):
            self.extractor_with_datas.\
                        get_datas_labeled_by_titles_alias(wrong_titles_alias)
