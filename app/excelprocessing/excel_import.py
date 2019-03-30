from openpyxl import load_workbook
from ..utils import list_contains_another
from openpyxl.utils.cell import column_index_from_string, get_column_interval,\
                                get_column_letter

class ExcelDataExtractor:

    def __init__(self, path):
        wb = load_workbook(path)
        self.ws = wb.active
        self.datas = None
        self.titles = None
        self.alias = None
        self.raw_datas = self.extract_raw_datas()
        self.init_datas_and_titles()

    def extract_raw_datas(self):
        raw_datas = []
        for row in self.ws.rows:
            row_values = self.get_row_values(row)
            if row_values is not None:
                raw_datas.append(row_values)
        return raw_datas

    def init_datas_and_titles(self):
        number_of_row_in_raw_datas = len(self.raw_datas)
        if number_of_row_in_raw_datas > 1:
            self.datas = self.raw_datas[1:]
            self.titles = self.raw_datas[0]
        elif number_of_row_in_raw_datas == 1:
            self.titles = self.raw_datas[0]

    def get_row_values(self, row):
        data = []
        is_empty = True
        for cell in row:
            data.append(cell.value)
            if cell.value is not None:
                is_empty = False
        if is_empty:
            return None
        else:
            return data

    def get_datas_labeled_by_titles_alias(self, titles_alias):
        aliased_keys_not_in_titles = not list_contains_another(self.titles,
                                                            titles_alias.keys())
        if aliased_keys_not_in_titles:
            raise KeyError("titles doesn't contain one or many aliased keys")
        return self.label_datas_by_titles_alias(titles_alias)

    def label_datas_by_titles_alias(self, titles_alias):
        columns_alias = {}
        for title in titles_alias:
            title_column_index = self.titles.index(title) + 1
            alias = titles_alias[title]
            columns_alias[title_column_index] = alias
        return self.get_datas_labeled_by_columns_alias(columns_alias)

    def get_datas_labeled_by_columns_alias(self, columns_alias):
        if self.datas is None:
            return None
        columns_alias_keys_not_in_columns_interval =\
                            not self.in_columns_interval(columns_alias.keys())
        if columns_alias_keys_not_in_columns_interval:
            raise KeyError('''interval from first to last no empty column
                              doesn't contain one or many aliased keys''')

        return self.label_datas_by_columns_alias(columns_alias)

    def in_columns_interval(self, columns_alias):
        min_column = 1
        max_column = len(self.titles)
        for key in columns_alias:
            if key < min_column or key > max_column:
                return False
        return True

    def label_datas_by_columns_alias(self, columns_alias):
        self.alias = columns_alias
        labeled_datas = []
        for data in self.datas:
            labeled_data = self.label_data_by_columns_alias(data)
            labeled_datas.append(labeled_data)
        return labeled_datas

    def label_data_by_columns_alias(self, data):
        labeled_data = {}
        for i in range(len(self.alias)):
           label_key = i + 1
           label = self.alias[label_key]
           labeled_data[label] = data[i]
        return labeled_data

    def get_datas(self):
        return self.datas

    def get_columns_titles(self):
        return self.titles
